# 160명이 제출한 과제파일을 학번으로 구분하여 각 실습강의실로 소분하는 프로그램

import shutil
import os
import openpyxl




def convert_original():
    # 수강생 전체 리스트가 담긴 엑셀파일 열기
    os.chdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus")
    excel_Total = openpyxl.load_workbook("02분반.xlsx")
    sheet_Total = excel_Total["Sheet1"]

    # 전체 리스트를 참조하여 분반 내용 딕셔너리로 만들기
    for rownum in range(1,200):
        # 엑셀파일 열기
        student_id = sheet_Total.cell(row = rownum, column = 5).value
        division = sheet_Total.cell(row = rownum, column = 2).value

        # 유효하지 않은 값 필터링
        if(student_id == None or division == None):
            continue

        # 분반 정보를 리스트로 저장
        if(division == 32423):
            division_32423.append(student_id)
        elif(division == 32530):
            division_32530.append(student_id)
        elif(division == 32531):
            division_32531.append(student_id)
        elif(division == 50313):
            division_50313.append(student_id)
        else:
            print("{}의 분류 정보가 없습니다. 학번을 확인해주세요. ".format(student_id))

    # 전체 학생 리스트를 생성.. 이는 나중에 미제출자 확인하기 위해서 만들어 두는 것
    student_list = division_32423 + division_32530 + division_32531 + division_50313

    return distribute(student_list)
    
def distribute(student_list):
    # 아이캠퍼스에서 다운로드한 파일이 위치한 주소로 이동
    os.chdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\Original")
    # 각각 분류할 폴더를 생성
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32423")
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32530")
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32531")
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\50313")
    
    # 총 제출 인원 및 분반별 제출 인원을 확인하기 위한 변수 선언
    count = [0,0,0,0,0,0] # 0은 총 제출 수, 1~4는 분반 기호, 5는 중복 제출자 수
                          # 따라서, 0에서 5를 빼면 총 제출 인원이 나옴
                    
    # Original 폴더에 있는 전체 목록을 하나씩 traversing 하면서 분류 시작
    total_list = os.listdir()
    # 분류가 완료된 리스트
    done = []
    for file in total_list:
        # 일단 그 파일이 무엇인지 출력
        print(file)
        # 더러운 파일 앞 10글자는 학번.. 그 부분만 슬라이싱해서 비교하기 위해 사용
        if(int(file[:10]) in division_32423):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32423")
            count[1] += 1
        elif(int(file[:10]) in division_32530):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32530")
            count[2] += 1
        elif(int(file[:10]) in division_32531):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32531")
            count[3] += 1
        elif(int(file[:10]) in division_50313):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\50313")
            count[4] += 1
        else:
            print("{}의 분류 정보가 없습니다. 학번을 확인해주세요. ".format(file[:10]))
        
        if(int(file[:10]) in done):
            # 여러파일을 제출한 학생 카운터 증가
            count[5] += 1            
        else:
            done.append(int(file[:10]))
            del student_list[student_list.index(int(file[:10]))]
        
        # 4개 분반의 제출 수를 더해서 총 제출수를 확인
        count[0] = count[1]+count[2]+count[3]+count[4]
    return (count, student_list)
    
def end(count, not_handin):
    print("총 {}개 과제 제출".format(count[0]))
    print("총 {}명 중 {}명 제출 {}명 미제출".format(165, count[0] - count[5], 165 - (count[0] - count[5])))

    print("미제출자 : ", end="")
    for student in not_handin:
        print("{} ".format(student), end="")
    print("")
    
    print("경제관 32423호 : {:3}개".format(count[1]))
    print("경제관 32530호 : {:3}개".format(count[2]))
    print("경제관 32531호 : {:3}개".format(count[3]))
    print("호암관 50313호 : {:3}개".format(count[4]))

# 구분할 빈 리스트 생성
division_32423 = []
division_32530 = []
division_32531 = []
division_50313 = []

# 전체 리스트에서 분반 리스트 생성

count = convert_original()
print("\n\n\n")
end(count[0], count[1])



