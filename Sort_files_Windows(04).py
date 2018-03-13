# 160명이 제출한 과제파일을 학번으로 구분하여 각 실습강의실로 소분하는 프로그램

import shutil
import os
import openpyxl




def convert_original():
    # 수강생 전체 리스트가 담긴 엑셀파일 열기
    os.chdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus")
    excel_Total = openpyxl.load_workbook("04분반.xlsx")
    sheet_Total = excel_Total["Sheet1"]

    # 전체 리스트를 참조하여 분반 내용 딕셔너리로 만들기


    for rownum in range(1,200):
        student_id = sheet_Total.cell(row = rownum, column = 5).value
        division = sheet_Total.cell(row = rownum, column = 2).value

        # 유효하지 않은 값 필터링
        if(student_id == None or division == None):
            continue

        # 분반 내역으로 
        if(division == 32423):
            division_32423.append(student_id)
            print("{}가 32423으로 분류되었습니다.".format(student_id))
        elif(division == 32530):
            division_32530.append(student_id)
            print("{}가 32530으로 분류되었습니다.".format(student_id))
        elif(division == 32531):
            division_32531.append(student_id)
            print("{}가 32531으로 분류되었습니다.".format(student_id))
        elif(division == 50313):
            division_50313.append(student_id)
            print("{}가 50313으로 분류되었습니다.".format(student_id))
        else:
            print("일단 오류...")

def distribute():
    os.chdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\Original")
    total_list = os.listdir()
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32423")
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32530")
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32531")
    os.mkdir("C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\50313")
    count = [0,0,0,0,0]
    for file in total_list:
        print(file)
        if(int(file[:10]) in division_32423):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32423")
            print("{}가 분류되었습니다.".format(file))
            count[1] += 1
        elif(int(file[:10]) in division_32530):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32530")
            print("{}가 분류되었습니다.".format(file))
            count[2] += 1
        elif(int(file[:10]) in division_32531):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\32531")
            print("{}가 분류되었습니다.".format(file))
            count[3] += 1
        elif(int(file[:10]) in division_50313):
            shutil.copy(file, "C:\\Users\\Joshua Y. S. Jung\\Desktop\\icampus\\50313")
            print("{}가 분류되었습니다.".format(file))
            count[4] += 1
        else:
            print("{} 오류..".format(file))

        count[0] = count[1]+count[2]+count[3]+count[4]
    return count
    
def end(count):
    print("총 제출 {}명 중 {}명".format(165, count[0]))
    print("32423 : {}명".format(count[1]))
    print("32530 : {}명".format(count[2]))
    print("32531 : {}명".format(count[3]))
    print("50313 : {}명".format(count[4]))

# 구분할 빈 리스트 생성
division_32423 = []
division_32530 = []
division_32531 = []
division_50313 = []

# 전체 리스트에서 분반 리스트 생성
convert_original()
print("\n\n\n")
count = distribute()
print("\n\n\n")
end(count)



